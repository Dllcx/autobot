# -*- coding: utf-8 -*-
"""
Created on Mon Apr 10 14:49:02 2017

@author: t684098
"""

import os
import time
import openpyxl
import datetime
import pandas as pd
import numpy as np
import requests
import win32com.client
from bs4 import BeautifulSoup
from bizdays import Calendar, load_holidays

class AutomatoUtilities(object):
    """
    Functions to suport all the others Automatos
    """
    @classmethod
    def di_crawler(cls):
        """
        Spider to bring DI from Cetip web page
        """
        page = requests.get("http://www.cetip.com.br")
        soup = BeautifulSoup(page.content, 'html.parser')
        di = soup.find(id="ctl00_Banner_lblTaxDI").text
        di = float(di[:2]+'.'+di[3:5])/100
        return di

    @classmethod
    def get_du(cls):
        """
        Create a business day calendar based on Ambima dates
        """
        holidays = load_holidays(
                r'C:\Users\t684098\Documents\Python\Feriados.txt')
        cal = Calendar(holidays, ['Sunday', 'Saturday'], name='Brazil')
        return cal
        
    @staticmethod
    def bizdays(ini_date,end_date):
        cal = AutomatoUtilities.get_du()
        return cal.bizdays(ini_date,end_date)
        
    @classmethod
    def gcb_column_picker(cls):
        """
        Pick the right column for the MiddleOffice GCB xls file
        """
        switcher_colummn = {1:3, 2:4, 3:5, 4:6, 5:7, 6:8, 7:9, 8:10, 9:11,
                            10:12, 11:13, 12:14}
        d0 = datetime.date.today() + datetime.timedelta(days=-2)
        if d0.weekday == 6:
            d = datetime.date.today() + datetime.timedelta(days=-4)
        elif d0.weekday == 5:
            d = datetime.date.today() + datetime.timedelta(days=-3)
        else:
            d = d0
        return switcher_colummn.get(d.month)
        
    @classmethod
    def corp_column_pick(cls):
        """
        Pick the right column for the MiddleOffice Corporate xls file
        """
        switcher_colummn = {1:3, 2:4, 3:5, 4:6, 5:7, 6:8, 7:9, 8:10, 9:11,
                            10:12, 11:13, 12:14}

        d0 = datetime.date.today() + datetime.timedelta(days=-2)
        if d0.weekday == 6:
            d = datetime.date.today() + datetime.timedelta(days=-4)
        elif d0.weekday == 5:
            d = datetime.date.today() + datetime.timedelta(days=-3)
        else:
            d = d0
        return switcher_colummn.get(d.month)
    
    @classmethod
    def inst_column_pick(cls):
        """
        Pick the right column for the MiddleOffice Institutionals xls file
        """
        switcher_colummn = {1:3, 2:4, 3:5, 4:6, 5:7, 6:8, 7:9, 8:10, 9:11,
                            10:12, 11:13, 12:14}
        d0 = datetime.date.today() + datetime.timedelta(days=-2)
        if d0.weekday == 6:
            d = datetime.date.today() + datetime.timedelta(days=-4)
        elif d0.weekday == 5:
            d = datetime.date.today() + datetime.timedelta(days=-3)
        else:
            d = d0
        return switcher_colummn.get(d.month)
    
    @classmethod
    def base_file_column_pick(cls):
        """
        Pick the right column on base xls file
        """
        # Set Date
        d = datetime.date.today() + datetime.timedelta(days=-2)
        # Column picker
        col_ref = 28
        if d.year == 2017:
            pass
        else:
            col_ref = col_ref + 12
        return col_ref + d.month

    @classmethod
    def updatecheck(cls,path):
        """
        Up to date verification
        """
        # Set Variables
        x = time.ctime(os.path.getmtime(path))
        y = time.strftime("%c")
        # Logic test and strng return
        if x[:10] == y[:10]:
            r_str = 0
        else:
            r_str = -1
        return r_str

class TaxAutomatoFront(object):
    """ 
    Automato Total Sheet module for tax update 
    """
    xls_base = r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média\Evolução Taxas Médias vs Finanças.xlsx'
    # Open the xls and get the correct sheet
    def __init__(self,ini_date,end_date):
        """
        Class initiator with dates to byzdays.
        Dates must be inserted as strings in 'yyyy-mm-dd' format
        """
        self.ini_date = ini_date
        self.end_date = end_date
        
    def total_xls_actions(self):
        """
        Total sheet update
        """
        # Open base file and pick Total sheet
        os.chdir(r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média')
        wb = openpyxl.load_workbook(TaxAutomatoFront.xls_base)
        sheet = wb.get_sheet_by_name('Total')
        # Get values for update
        col = AutomatoUtilities.base_file_column_pick()
        di = AutomatoUtilities.di_crawler()
        bdays = AutomatoUtilities.bizdays(self.ini_date, self.end_date)
        # Cells update
        sheet.cell(row=1, column=col).value = di
        sheet.cell(row=2, column=col).value = bdays
        # wb save and close
        wb.save('Evolução Taxas Médias vs Finanças.xlsx')  
        wb.close()
        return "Total Sheet update complete"

class TaxAutomatoGCB(object):
    """ 
    Automato GCB Corp Sales module for tax update 
    """
    xls_base = r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média\Evolução Taxas Médias vs Finanças.xlsx'
    def __init__(self,ini_date,end_date):
        """
        Class initiator with dates to byzdays. 
        Dates must be inserted as strings in 'yyyy-mm-dd' format
        """
        self.ini_date = ini_date
        self.end_date = end_date
        self.path = r'\\bsbrsp369\Treasury\Sales\GB&M Derivativos\Resultados\Rates_Analitico_GCB.xlsm'
    
    def gcb_open_remotexls(self):
        """
        GCB Corp Sales sheet update remote files values
        """
        gcb_vals = []
        print ("Arquivo remoto atualizado")
        # Open remote xls file and get GCB vals
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name('analise_RF')
        col = AutomatoUtilities.gcb_column_picker()
        # DAP avrg O/S
        gcb_vals.append(sheet.cell(row=8, column=col).value)
        # DAP revenue
        gcb_vals.append(sheet.cell(row=18, column=col).value)
        # Comp. avrg O/S
        gcb_vals.append(sheet.cell(row=9, column=col).value)
        # Comp. revenue
        gcb_vals.append(sheet.cell(row=19, column=col).value)        
        wb.close()
        return gcb_vals
    
    def gcb_get_average(self):
        """
        GCB Corp Sales sheet update remote files average
        """
        avrg_list = []
        xl = pd.ExcelFile(self.path)
        df = xl.parse("bd_renda_fixa")
        d = datetime.date.today() + datetime.timedelta(days=-2)
        if d.month/2 < 5:
            yfilter = str(d.year) + '0' + str(d.month)
        else:
            yfilter = str(d.year) + str(d.month)
        df = df[df['Ano/Mes'] == int(yfilter)]
        # Comp mean value
        df_comp = df[df['DESCRIÇÃO BP'] == 'Operações Compromissadas']
        df_comp = df_comp[df_comp['Taxa_Captacao'] != 0]
        avr_comp = np.average(df_comp['Taxa_Captacao'], 
                              weights=df_comp['SALDO MÉDIO'])
        # DAP mean value
        df_DAP = df[df['DESCRIÇÃO BP'] == 'Depósito a Prazo']
        df_DAP = df_DAP[df_DAP['Taxa_Captacao'] != 0]
        avr_DAP = np.average(df_DAP['Taxa_Captacao'], 
                             weights=df_DAP['SALDO MÉDIO'])
        avrg_list.append(avr_comp / 100)
        avrg_list.append(avr_DAP / 100)
        return avrg_list
    
    def gcb_xls_actions(self):
        """
        Update and values replacement in "GCB CORP SALES" sheet
        """
        # Verify remote file update
        check_vlw = AutomatoUtilities.updatecheck(self.path)
        if check_vlw == 0:
            os.chdir(r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média')
            wb = openpyxl.load_workbook(TaxAutomatoGCB.xls_base)
            sheet = wb.get_sheet_by_name('GCB Corp Sales')
            col = AutomatoUtilities.base_file_column_pick()
            # Average picker
            avr_lst = TaxAutomatoGCB.gcb_get_average(self)
            # Values Replacement
            sheet.cell(row=3, column=col).value = avr_lst[0]
            sheet.cell(row=6, column=col).value = avr_lst[1]
            # O/S and Revnues list pick
            xls_lst = TaxAutomatoGCB.gcb_open_remotexls(self)
            # Values Replacement
            sheet.cell(row=10, column=col).value = xls_lst[2]
            sheet.cell(row=12, column=col).value = xls_lst[0]
            # Date Set up
            cal = AutomatoUtilities.get_du()
            d0 = datetime.date.today() + datetime.timedelta(days=-2)
            if d0.weekday == 6:
                d = datetime.date.today() + datetime.timedelta(days=-4)
            elif d0.weekday == 5:
                d = datetime.date.today() + datetime.timedelta(days=-3)
            else:
                d = d0
            # Revenues Replacement
            if d == cal.getdate('last bizday', d.year, d.month):
                sheet.cell(row=11, column=col).value = xls_lst[3]
                sheet.cell(row=13, column=col).value = xls_lst[1]
            else:
                month = str(d.month)
                if d.month / 2 < 5:
                    month_ = '0' + month
                else:
                    month_ = month
                
                if d.day / 2 < 5:
                    day2_ = '0' + str(d.day)
                else:
                    day2_ = str(d.day)
                
                update_date = str(d.year) + '-' + month_ + '-' + day2_
                bdays = cal.bizdays(self.ini_date, update_date)
                full_bdays = AutomatoUtilities.bizdays(self.ini_date,self.end_date)
                sheet.cell(row=11, column=col).value=(xls_lst[3]/bdays)*full_bdays
                sheet.cell(row=13, column=col).value=(xls_lst[1]/bdays)*full_bdays
            wb.save('Evolução Taxas Médias vs Finanças.xlsx')
            wb.close()
            rt_val = 1
        else:
            rt_val = 0
        return rt_val
        
class TaxAutomatoInst(object):
    """ 
    Automato GCB Inst module for tax update 
    """
    xls_base = r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média\Evolução Taxas Médias vs Finanças.xlsx'

    def __init__(self,ini_date,end_date):
        """
        Class initiator with dates to byzdays. 
        Dates must be inserted as strings in 'yyyy-mm-dd' format
        """
        self.ini_date = ini_date
        self.end_date = end_date
        self.path = r'\\bsbrsp369\Treasury\Distrib\Rates_Analitico_Inst.xlsm'

    def inst_open_remotexls(self):
        """
        GCB Inst Sales sheet update remote files values
        """
        inst_vals = []
        print ("Arquivo remoto atualizado")
        # Open remote xls file and get GCB vals
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name('analise_RF')
        col = AutomatoUtilities.inst_column_pick()
        # DAP avrg O/S
        inst_vals.append(sheet.cell(row=8, column=col).value)
        # DAP revenue
        inst_vals.append(sheet.cell(row=13, column=col).value)
        # Comp. avrg O/S
        inst_vals.append(sheet.cell(row=9, column=col).value)
        # Comp. revenue
        inst_vals.append(sheet.cell(row=14, column=col).value)        
        wb.close()
        return inst_vals
        
    def inst_get_average(self):
        """
        GCB Inst Sales sheet update remote files average
        """
        avrg_list = []
        xl = pd.ExcelFile(self.path)
        df = xl.parse("bd_renda_fixa")
        d = datetime.date.today() + datetime.timedelta(days=-2)
        if d.month/2 < 5:
            yfilter = str(d.year) + '0' + str(d.month)
        else:
            yfilter = str(d.year) + str(d.month)
        df = df[df['Ano/Mes'] == int(yfilter)]
        # Comp mean value
        df_comp = df[df['DESCRIÇÃO BP'] == 'Operações Compromissadas']
        df_comp = df_comp[df_comp['Taxa_Captacao'] != 0]
        avr_comp = np.average(df_comp['Taxa_Captacao'], 
                                  weights=df_comp['SALDO MÉDIO'])
        # DAP mean value
        df_DAP = df[df['DESCRIÇÃO BP'] == 'Depósito a Prazo']
        df_DAP = df_DAP[df_DAP['Taxa_Captacao'] != 0]
        avr_DAP = np.average(df_DAP['Taxa_Captacao'], 
                                 weights=df_DAP['SALDO MÉDIO'])
        avrg_list.append(avr_comp / 100)
        avrg_list.append(avr_DAP / 100)
        return avrg_list

    def inst_xls_actions(self):
        """
        Update and values replacement in "GCB Inst" sheet
        """
        # Verify remote file update
        check_vlw = AutomatoUtilities.updatecheck(self.path)
        if check_vlw == 0:
            os.chdir(r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média')
            wb = openpyxl.load_workbook(TaxAutomatoInst.xls_base)
            sheet = wb.get_sheet_by_name('GCB Inst')
            col = AutomatoUtilities.base_file_column_pick()
            # Average picker
            avr_lst = TaxAutomatoInst.inst_get_average(self)
            # Values Replacement
            sheet.cell(row=3, column=col).value = avr_lst[0]
            sheet.cell(row=6, column=col).value = avr_lst[1]
            # O/S and Revnues list pick
            xls_lst = TaxAutomatoInst.inst_open_remotexls(self)
            # Values Replacement
            sheet.cell(row=10, column=col).value = xls_lst[2]
            sheet.cell(row=12, column=col).value = xls_lst[0]
            # Date set up
            cal = AutomatoUtilities.get_du()
            d0 = datetime.date.today() + datetime.timedelta(days=-2)
            if d0.weekday == 6:
                d = datetime.date.today() + datetime.timedelta(days=-4)
            elif d0.weekday == 5:
                d = datetime.date.today() + datetime.timedelta(days=-3)
            else:
                d = d0
            # Revenues Replacement
            if d == cal.getdate('last bizday', d.year, d.month):
                sheet.cell(row=11, column=col).value = xls_lst[3]
                sheet.cell(row=13, column=col).value = xls_lst[1]
            else:
                month = str(d.month)
                if d.month / 2 < 5:
                    month_ = '0' + month
                else:
                    month_ = month
                
                if d.day / 2 < 5:
                    day2_ = '0' + str(d.day)
                else:
                    day2_ = str(d.day)
                
                update_date = str(d.year) + '-' + month_ + '-' + day2_
                bdays = cal.bizdays(self.ini_date, update_date)
                full_bdays = AutomatoUtilities.bizdays(self.ini_date,self.end_date)
                sheet.cell(row=11, column=col).value=(xls_lst[3]/bdays)*full_bdays
                sheet.cell(row=13, column=col).value=(xls_lst[1]/bdays)*full_bdays
            wb.save('Evolução Taxas Médias vs Finanças.xlsx')
            wb.close()
            rt_val = 1
        else:
            rt_val = 0
        return rt_val

class TaxAutomatoCorp(object):
    """ 
    Automato Corporate module for tax update 
    """
    xls_base = r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média\Evolução Taxas Médias vs Finanças.xlsx'

    def __init__(self,ini_date,end_date):
        """
        Class initiator with dates to byzdays. 
        Dates must be inserted as strings in 'yyyy-mm-dd' format
        """
        self.ini_date = ini_date
        self.end_date = end_date
        self.path = r'\\bsbrsp369\Treasury\Sales\Mesa Empresas\MO\Rates_Analitico_Corporate.xlsm'
        self.path_avrg = r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média\Evolução Taxas Média - access MO.xlsx'

    def corp_open_remotexls(self):
        """
        Corp Sales sheet update remote files values
        """
        corp_vals = []
        print ("Arquivo remoto atualizado")
        # Open remote xls file and get GCB vals
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name('analise_RF')
        col = AutomatoUtilities.inst_column_pick()
        # DAP avrg O/S
        corp_vals.append(sheet.cell(row=8, column=col).value)
        # DAP revenue
        corp_vals.append(sheet.cell(row=21, column=col).value)
        # Comp. avrg O/S
        corp_vals.append(sheet.cell(row=9, column=col).value)
        # Comp. revenue
        corp_vals.append(sheet.cell(row=22, column=col).value)        
        wb.close()
        return corp_vals
        
    def corp_get_average(self):
        """
        Corporate Inst Sales sheet update remote files average
        """
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        wb = xl.Workbooks.Open(self.path_avrg)
        ws = wb.Worksheets('Sheet1')
        wb.RefreshAll()
        time.sleep(200)
        ws = wb.Worksheets('Sheet4')
        ws.PivotTables(1).PivotCache().Refresh()
        # Base Columns set up
        col_CDB = 7
        col_OC = 8
        rw_ref = 5
        # Date logic resolution
        d0 = datetime.date.today() + datetime.timedelta(days=-2)
        if d0.weekday == 6:
            d = datetime.date.today() + datetime.timedelta(days=-4)
        elif d0.weekday == 5:
            d = datetime.date.today() + datetime.timedelta(days=-3)
        else:
            d = d0
        # Row Definition
        rw = rw_ref + d.month
        # Values vector
        corp_avrg = []
        corp_avrg.append(ws.cells(rw,col_OC).value / 100)
        corp_avrg.append(ws.cells(rw,col_CDB).value / 100)
        wb.Close(True)
        xl.Quit()
        return corp_avrg

    def corp_xls_actions(self):
        """
        Update and values replacement in "GCB Inst" sheet
        """
        # Verify remote file update
        check_vlw = AutomatoUtilities.updatecheck(self.path)
        if check_vlw == 0:
            os.chdir(r'\\mscluster11fs\SMI\Ativos - Investimentos\Evolução Volume Taxa Média')
            wb = openpyxl.load_workbook(TaxAutomatoInst.xls_base)
            sheet = wb.get_sheet_by_name('Corporate')
            col = AutomatoUtilities.base_file_column_pick()
            # Average picker
            avr_lst = TaxAutomatoCorp.corp_get_average(self)
            # Values Replacement
            sheet.cell(row=3, column=col).value = avr_lst[0]
            sheet.cell(row=6, column=col).value = avr_lst[1]
            # O/S and Revnues list pick
            xls_lst = TaxAutomatoCorp.corp_open_remotexls(self)
            # Values Replacement
            sheet.cell(row=10, column=col).value = xls_lst[2]
            sheet.cell(row=12, column=col).value = xls_lst[0]
            # Date set up
            cal = AutomatoUtilities.get_du()
            d0 = datetime.date.today() + datetime.timedelta(days=-2)
            if d0.weekday == 6:
                d = datetime.date.today() + datetime.timedelta(days=-4)
            elif d0.weekday == 5:
                d = datetime.date.today() + datetime.timedelta(days=-3)
            else:
                d = d0
            # Revenues Replacement
            if d == cal.getdate('last bizday', d.year, d.month):
                sheet.cell(row=11, column=col).value = xls_lst[3]
                sheet.cell(row=13, column=col).value = xls_lst[1]
            else:
                month = str(d.month)
                if d.month / 2 < 5:
                    month_ = '0' + month
                else:
                    month_ = month
                
                if d.day / 2 < 5:
                    day2_ = '0' + str(d.day)
                else:
                    day2_ = str(d.day)
                
                update_date = str(d.year) + '-' + month_ + '-' + day2_
                bdays = cal.bizdays(self.ini_date, update_date)
                full_bdays = AutomatoUtilities.bizdays(self.ini_date,self.end_date)
                sheet.cell(row=11, column=col).value=(xls_lst[3]/bdays)*full_bdays
                sheet.cell(row=13, column=col).value=(xls_lst[1]/bdays)*full_bdays
            wb.save('Evolução Taxas Médias vs Finanças.xlsx')
            wb.close()
            rt_val = 1
        else:
            rt_val = 0
        return rt_val